#!/usr/bin/env python
# coding=utf-8

from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkiot.request.v20180120.QueryDevicePropertyStatusRequest import (
    QueryDevicePropertyStatusRequest,
)
from time_transform import timestamp_to_format, timeformat_to_timestamp
import datetime
import time
import json
import openpyxl

# 定义表格操作全局变量
n_Distance = 0  # Distance值
n_amp = 0  # amp值
n_valu = 0  # valu


def Realtime_data(ac, pp):
    client = AcsClient(
        "LTAI4FbnNz12HT85LUPbrFnU", "UdPN3kNhB8aVJvrrbfIyng5sHupNpA", "cn-shanghai"
    )

    request = QueryDevicePropertyStatusRequest()
    request.set_accept_format("json")

    # request.set_IotId("IotId")
    request.set_ProductKey(pp)
    request.set_DeviceName(ac)

    response = client.do_action_with_exception(request)
    # python_amp:  print(response)
    # print(response)
    byte = response
    str1 = byte.decode()
    json_str = json.loads(str1)
    # print(json_str)
    a = json_str["Data"]
    # print(a)
    # print("\n")
    b = a["List"]
    # print(b,'\n')
    c = b["PropertyStatusInfo"]
    # print(c)
    # d=json.loads(c)
    au = {}
    for i in range(0, len(c)):
        # print(c[i]['Identifier'])
        ad = c[i]["Identifier"]
        # c[i]['Identifier']= c[i]['Identifier']
        if (
            c[i]["Identifier"] == "electricity"
            or c[i]["Identifier"] == "Distance"
            or c[i]["Identifier"] == "Distance1"
            or c[i]["Identifier"] == "Distance2"
            or c[i]["Identifier"] == "amp"
            or c[i]["Identifier"] == "valu"
            or c[i]["Identifier"] == "temp"
        ):
            if c[i]["Identifier"] == "Distance":
                # print(timestamp_to_format(int(c[i]['Time'])/1000))
                au["Time"] = timestamp_to_format(int(c[i]["Time"]) / 1000)
            # print(c[i]['Value'], timestamp_to_format(int(c[i]['Time'])/1000))
            ap = c[i]["Value"]
            au[ad] = ap
            au["井号"] = ac
    # print(au)
    return au

    # if __name__ == '__main__':
def query(li):
        # 选择要查询的区域
    print(
        "查询郑庄采油队张台项目区请输入1",
        "\n",
        "郭旗采油队羊圈山采油站请输入2",
        "\n",
        "黑家堡采油队贺家沟采油站请输入3",
        "\n",
        "黑家堡采油队王槐子沟采油站请输入4",
    )
    # li = input("请选择要查询的区域：")
    print(li)
    # 初始化表格
    title1 = datetime.datetime.now().strftime("%m%d%H%M%S")
    str_time = "TL" + str(title1)  # 拼接新建表名
    print("创建", str_time, "表格成功")
    # shell = openpyxl.Workbook()
    shell = openpyxl.load_workbook("实时数据.xlsx")
    b1 = shell.create_sheet(str_time, 0)
    b1.column_dimensions["B"].width = 20
    b1.column_dimensions["I"].width = 12
    b1.cell(row=1, column=1, value="井号")
    b1.cell(row=1, column=2, value="罐号")
    b1.cell(row=1, column=3, value="Distance")
    b1.cell(row=1, column=4, value="Distance1")
    b1.cell(row=1, column=5, value="Distance2")
    b1.cell(row=1, column=6, value="valu")
    b1.cell(row=1, column=7, value="amp")
    b1.cell(row=1, column=8, value="temp")
    b1.cell(row=1, column=9, value="electricity")
    # 建立查询列表(局部变量)
    # 张台

    ct = [
        "YAN39",
        "YAN39-1",
        "YAN39-3",
        "YAN39-5",
        "YAN39-6",
        "YAN76",
        "YAN76-3",
        "YAN76-5",
        "YAN76-7",
        "YAN79-1",
        "YAN79-2",
        "YAN79-3",
        "YAN79-4",
        "YAN79-5",
        "YAN79-6",
        "YAN79-7",
        "YAN79-8",
        "Z671-1",
        "Z671-2",
        "Z671-3",
        "Z671-4",
        "Z671-5",
        "Z671-6",
        "Z671-7",
        "Z671-8",
        "Z672",
        "Z672-3",
        "Z673",
        "Z673-1",
        "Z673-2",
        "Z673-7",
        "Z674-1",
        "Z674-3",
        "Z674-5",
        "Z674-7",
        "Z674-8",
        "Z676",
        "Z676-4",
        "Z676-6",
        "Z676-7",
        "Z677-1",
        "Z677-2",
        "Z677-33",
        "Z677-4",
        "Z677-5",
        "Z677-6",
        "Z677-7",
        "Z677-8",
        "Z678",
        "Z678-6",
        "Z678-7",
        "Z679-1",
        "Z679-2",
        "Z679-4",
        "Z679-5",
        "Z679-6",
        "Z679-8",
        "Z683-1",
        "Z683-2",
        "Z683-4",
        "Z683-5",
        "Z683-6",
        "Z683-7",
        "Z683-8",
        "Z687",
        "Z687-1",
        "Z687-5",
        "Z687-7",
        "Z688-7",
        "Z688-8",
        "Z688-9",
        "Z703-1",
        "Z703-7",
        "Z703-8",
        "Z862-4",
        "Z862-6",
        "Z862-8",
        "Z862-9",
    ]
    # 郭旗
    ct1 = [
        "GUO546-5",
        "GUO546-4",
        "GUO546-3",
        "GUO546-2",
        "GUO546",
        "YAN56-8",
        "YAN56-7",
        "YAN56-6",
        "YAN56-5",
        "YAN56-4",
        "YAN56-3",
        "YAN56-2",
        "YAN56-1",
        "GUO600-7",
        "GUO600-6",
        "GUO600-3",
        "GUO600-2",
        "GUO600",
        "GUO601-7",
        "GUO601-3",
        "GUO601",
    ]
    # 二区
    ct2 = [
        "Q206",
        "Q1500",
        "Q1501",
        "Q1502",
        "Q1504",
        "Q1505",
        "Q1506",
        "Q1507",
        "Q1508",
        "Q1509",
        "Q1054",
        "Q1329-3",
        "Q1329-4",
        "Q1329",
        "Q1329-7",
        "Q1329-8",
        "Q1328-4",
        "Q1328-2",
        "Q1328-6",
        "Q1328-8",
        "Q994",
        "Q995",
        "Q998",
        "Q1380",
        "Q1381",
        "Q1382",
        "Q1383",
        "Q1384",
        "Q1385",
        "Q1386",
        "Q1387",
        "Q1388",
        "Q1389",
        "Q1390",
        "Q1391",
        "Q1392",
        "Q1393",
        "Q1394",
        "Q1395",
        "Q1396",
        "Q1397",
        "Q1398",
        "Q1399",
        "Q1400",
        "Q1401",
        "Q1402",
        "Q1403",
        "Q1404",
        "Q1405",
        "Q1406",
        "Q1407",
        "Q1408",
        "Q1409",
        "Q1410",
        "Q1411",
        "Q1412",
        "Q1413",
        "Q1414",
        "Q1415",
        "Q1416",
        "Q1417",
        "Q1418",
        "Q1419",
        "Q959",
        "Q1083",
        "Q1084",
        "Q1089",
        "Q1091",
        "Q1189",
        "Q1190",
        "Q1191",
        "Q1192",
        "Q1193",
        "Q1194",
        "Q1195",
        "Q1196",
        "Q1197",
        "Q1198",
        "Q922",
        "Q923",
        "Q931",
        "Q932",
        "Q933",
        "Q934",
        "Q936",
        "Q938",
        "Q939",
        "Q940",
        "Q948",
        "Q950",
        "Q951",
        "Q1364",
        "Q1365",
        "Q1366",
        "Q1308",
        "Q1309",
        "Q1310",
        "Q1311",
        "Q1312",
        "Q1368",
        "Q1369",
        "Q1370",
        "Q1371",
        "Q1372",
        "Q1373",
        "Q1374",
        "Q1375",
        "Q1376",
        "Q1377",
        "Q1379",
        "Yan35",
        "Q949",
        "Q1030",
        "Q1031",
        "Q1037",
        "Q1162",
        "Q1163",
        "Q1164",
        "Q1165",
        "Q1166",
        "Q1167",
        "Q1168",
        "Q1169",
        "Q1170",
        "Q1171",
        "Q1172",
        "Q1173",
        "Q1174",
        "Q1175",
        "Q1176",
        "Q1178",
        "Q1179",
        "Q1180",
        "Q1181",
        "Q1182",
        "Q1183",
        "Q1184",
        "Q1185",
        "Q1186",
        "Q1187",
        "Q1188",
        "Q1017",
        "Q1018",
        "Q1024",
        "Q1025",
        "Q1026",
        "Q1027",
        "Q1029",
        "Q1033",
        "Q1034",
        "Q1043",
        "Q1044",
        "Q1045",
        "Q1046",
        "Q1047",
        "Q1048",
        "Q1049",
        "Q955",
        "Q956",
        "Q957",
        "Q1301",
        "Q944",
        "Q945",
        "Q958",
        "Q941",
        "Q942",
        "Q1340",
        "Q1341",
        "Q1342",
        "Q1343",
        "Q1344",
        "Q1345",
        "Q1346",
        "Q1349",
        "Q1350",
        "Q1351",
        "Q1353",
        "Q1354",
        "Q1355",
        "Q1356",
        "Q1357",
        "Q946",
        "Q947",
        "Q1292",
        "Q1293",
        "Q1294",
        "Q1296",
        "Q1297",
        "Q953",
        "Q954",
        "Q1361",
        "Q1362",
        "Q1363",
        "Q975",
        "Q1269",
        "Q1270",
        "Q1032",
    ]
    # 三区
    ct3 = [
        "G117",
        "Z017",
        "G125",
        "Z049",
        "G158",
        "G158P-2NO.1",
        "G158P-2NO.2",
        "G158P-2NO.3",
        "G158P-3NO.1",
        "G158P-3NO.2",
        "G158P-3NO.3",
        "G158P-6NO.1",
        "G158P-6NO.2",
        "G158P-6NO.3",
        "G170",
        "G170-2",
        "G170-3",
        "G170-4",
        "G170-6",
        "G170-7",
        "G170-8",
        "G124",
        "G135",
        "G135-2",
        "G135-3",
        "G135-4",
        "G135-6",
        "G135-7",
        "G135-8",
        "G185",
        "G185-2",
        "G185-3",
        "G185-7",
        "G185-8",
        "G142",
        "G226P4No.1",
        "G226P4No.2",
        "G226P6No.1",
        "G226P6No.2",
        "Z015",
        "Z074",
        "Z074-2",
        "Z074-3",
        "Z074-4",
        "Z074-6",
        "Z074-7",
        "Z074-8",
        "G122",
        "G123",
        "G126-2",
        "G126-4",
        "G126-6",
        "G126-8",
        "G127-2",
        "G127-4",
        "G127-6",
        "G127-8",
        "G129",
        "G129-2",
        "G129-3",
        "G129-4",
        "G129-6",
        "G129-7",
        "G129-8",
        "Z052",
        "Z052-2",
        "Z052-3",
        "Z052-4",
        "Z052-6",
        "Z052-7",
        "Z052-8",
        "G202",
        "G202-2",
        "G202-4",
        "G202-5",
        "G202-6",
        "G202-8",
        "Z034",
        "Z034-2",
        "Z034-3",
        "Z034-4",
        "Z034-6",
        "Z034-7",
        "Z034-8",
        "G207",
        "G207-1",
        "G207-2",
        "G207-4",
        "G207-5",
        "G207-6",
        "G207-8",
        "G136",
        "G136-2",
        "G136-3",
        "G136-4",
        "G136-6",
        "G136-7",
        "G136-8",
        "Z051",
        "G187",
        "G187-2",
        "G187-3",
        "G187-4",
        "G187-6",
        "G187-7",
        "G187-8",
        "G118",
        "Z037",
        "G206",
        "G206-1",
        "G206-3",
        "G206-5",
        "G206-6",
        "G206-8",
        "Z053",
        "Z053-2",
        "Z053-3",
        "Z053-4",
        "Z053-6",
        "G120",
        "G121",
        "Z046",
        "Z062",
        "Z062-2",
        "Z062-4",
        "Z062-6",
        "Z062-7",
        "G119",
        "G133-2",
        "G133-4",
        "G133-6",
        "G133-8",
        "Z019",
        "Z019-2",
        "Z019-3",
        "Z019-4",
        "Z019-6",
        "Z019-7",
        "Z019-8",
        "Z036",
        "Z035",
        "G113",
        "Z030",
        "Z054",
        "G115",
        "Z033",
        "Y110",
        "Z018",
        "Z055",
        "G108",
        "G109",
        "Yan96",
        "G112",
        "G146-2",
        "G146-4",
        "G146-6",
        "G146-8",
        "G148-2",
        "G148-4",
        "G148-6",
        "Yan67",
        "G111",
        "G150-2",
        "G150-8",
        "G165-2",
        "G165-4",
        "G165-6",
        "G165-8",
        "G167-2",
        "G167-4",
        "G167-6",
        "G167-8",
        "G153-6",
        "G153-8",
        "Z064",
        "G110",
        "Yan66",
        "G151-2",
        "G151-4",
        "G151-6",
        "G151-8",
        "G102",
        "G103",
        "G105",
        "G106",
        "Yan65",
        "Z069",
        "Yan63",
        "Yan64",
        "Yan68",
        "Yan69",
        "G104",
        "Q1259",
        "Q1260",
        "Q1261",
        "Q1262",
        "Q1264",
        "Q1265",
        "Q1266",
        "Q1267",
        "Q1268",
        "G116",
        "Z062-3",
        "Z062-8",
        "Z053-8",
        "Z052P-8NO.1",
        "Z052P-8NO.2",
        "Q10P-3NO.1",
        "Q10P-3NO.2",
    ]

    cp = [
        "井号",
        "Time",
        "Distance",
        "Distance1",
        "Distance2",
        "valu",
        "amp",
        "temp",
        "electricity",
    ]
    list_device = [ct, ct1, ct2, ct3]  # 可调用区域列表

    list_product = ["a1gMuOnqrr7", "a1Mag4tNOD3", "a1g1PqlfkOo", "a1HyPOJ4feR"]
    # 写表格操作
    li = int(li) - 1
    for k in range(0, len(list_device[li])):  # 列循环
        print(list_device[li][k])
        au = Realtime_data(list_device[li][k], list_product[li])  # 获取查询结果
        for r in range(2, 11):  # 行循环
            xp = cp[r - 2]  # 查询到的字典键
            if (
                cp[r - 2] == "Distance"
                or cp[r - 2] == "Distance1"
                or cp[r - 2] == "Distance2"
                or cp[r - 2] == "temp"
                or cp[r - 2] == "electricity"
            ):
                n_Distance = float(au[xp])  # 查询到的字典值
                if cp[r - 2] == "Distance":  # 对Distance的值操作
                    if n_Distance > 80:
                        print("罐存过高，请及时组织泵出")
                        red_fill = openpyxl.styles.PatternFill(
                            "solid", fgColor="1AFD9C"
                        )
                        b1.cell(k + 2, r - 1).fill = red_fill
                        if n_Distance > 150:
                            print("传感器异常")
                            red_fill = openpyxl.styles.PatternFill(
                                "solid", fgColor="FF5809"
                            )
                            b1.cell(k + 2, r - 1).fill = red_fill
                    elif n_Distance < 0:
                        print("传感器异常")
                        red_fill = openpyxl.styles.PatternFill(
                            "solid", fgColor="FF5809"
                        )
                        b1.cell(k + 2, r - 1).fill = red_fill
                b1.cell(row=k + 2, column=r - 1, value=n_Distance)
            elif cp[r - 2] == "amp":  # 对amp的值操作
                n_amp = int(au[xp])
                if au[xp] == None:
                    au[xp] = 0
                b1.cell(row=k + 2, column=r - 1, value=n_amp)
            elif cp[r - 2] == "井号" or cp[r - 2] == "Time":
                if cp[r - 2] == "Time":  # 对上线时间的操作
                    n0 = au[xp]
                    nx = timeformat_to_timestamp(n0)
                    n_amp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ny = timeformat_to_timestamp(n_amp)
                    xiaojiu = ny - nx
                    if xiaojiu > 86400:  # 1153021   一天86400000毫秒
                        jiujiu = xiaojiu / 86400
                        print("\b", "设备", "{:.2f}".format(jiujiu), "天内未上线")
                        red_fill = openpyxl.styles.PatternFill(
                            "solid", fgColor="ff7575"
                        )
                        b1.cell(k + 2, r - 1).fill = red_fill
                    b1.cell(row=k + 2, column=r - 1, value=au[xp])
                b1.cell(row=k + 2, column=r - 1, value=au[xp])
            elif cp[r - 2] == "valu":  # 对校验值的操作
                n_valu = float(au[xp])
                if n_Distance == n_valu:
                    print("传感器异常")
                    red_fill = openpyxl.styles.PatternFill(
                        "solid", fgColor="FF5809"
                    )
                    b1.cell(k + 2, r - 1).fill = red_fill
                b1.cell(row=k + 2, column=r - 1, value=n_valu)
    shet = shell.sheetnames  # 显示已存在表名，该对象为列表
    while len(shet) >= 15:
        shut = shet[-1]
        ws = shell[shut]  # 表名赋值为删除对象
        shell.remove(ws)
        shet.remove(shet[-1])
        print("早期表格清理成功")
    shell.save("实时数据.xlsx")
    print("程序执行完成")
    # input("等待输入指令")
